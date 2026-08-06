"""
Parser for Vinmonopolet historical LAUNCH LISTS (data-from-vinmonopolet/*.xlsx).

These are NOT tender requests (what VMP asks for) — they are what actually got LAUNCHED:
one row per listed product, with producer, origin (down to vineyard), classification, type,
vintage, ABV, the real retail price, total quantity + value, per-store allocation, and — on the
spirits lists — the importer (Grossist) and distributor. A real 6-year market history.

Consolidates every file into launch_history.json (one clean record per product per launch).

Usage:  python3 parse_launch_list.py [--out launch_history.json]
"""
import glob, json, os, re, sys, unicodedata, warnings
warnings.filterwarnings("ignore")
from openpyxl import load_workbook

DATA_DIR = "data-from-vinmonopolet"


def norm(s):
    if s is None:
        return ""
    s = str(s).translate(str.maketrans({"ø": "o", "Ø": "O", "æ": "ae", "Æ": "Ae", "å": "a", "Å": "A"}))
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


# header (normalized) -> canonical field
HEAD = {
    "artikkelnr.": "art", "artikkelnr": "art", "varenummer": "art",
    "produsent": "producer", "produktnavn": "product",
    "argang": "vintage", "land": "country",
    "distrikt": "district", "region": "district", "kommune": "municipality",
    "vinmark": "vineyard", "klassifikasjon": "classification", "klassifisering": "classification",
    "varetype": "type", "stil": "type",
    "volum": "volume", "alkohol": "abv", "salgspris": "price",
    "grossist": "importer", "distributor": "distributor",
    "totalt antall": "qty", "antall kjop": "qty", "totalverdi": "value",
}

TYPE_CANON = [
    (r"rodvin", "red"), (r"hvitvin", "white"), (r"rosevin|rose", "rose"),
    (r"musser", "sparkling"), (r"sterkvin|hetvin|portvin|sherry|madeira", "fortified"),
    (r"whisky|whiskey", "whisky"), (r"rom\b|rum", "rum"), (r"gin\b", "gin"),
    (r"cognac|konjakk|armagnac|brandy", "brandy"), (r"tequila|mezcal|raicilla", "agave"),
    (r"akevitt|aquavit", "aquavit"), (r"vodka", "vodka"), (r"likor|liqueur", "liqueur"),
    (r"brennevin", "spirits-other"),
    (r"pale ale|ipa|porter|stout|saison|surol|lager|pils|klosterstil|spesial|barley|ol\b", "beer"),
    (r"sider|cider", "cider"),
]


def canon_type(t):
    n = norm(t)
    for pat, lab in TYPE_CANON:
        if re.search(pat, n):
            return lab
    return n or ""


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(v))
    return float(m.group(0).replace(",", ".")) if m else None


def period_and_category(fname):
    base = os.path.basename(fname)
    m = re.search(r"(\d{4})(\d{2})", base)
    period = f"{m.group(1)}-{m.group(2)}" if m else "????-??"
    # category hint = filename minus the date, the word "lanseringsliste", and boilerplate
    cat = re.sub(r"\.xlsx$", "", base, flags=re.I)
    cat = re.sub(r"\d{6}", " ", cat)
    cat = re.sub(r"lanserings?liste|vinmonopolet|liste|\+|_|-", " ", cat, flags=re.I)
    cat = re.sub(r"\b(januar|februar|mars|april|mai|juni|juli|august|september|oktober|november|desember)\b",
                 " ", cat, flags=re.I)
    cat = re.sub(r"\b\d+\.?\b|\bref\d+\b", " ", cat)
    cat = re.sub(r"\s+", " ", cat).strip()
    return period, (cat or "generell")


def parse_file(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    period, category = period_and_category(path)
    out = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        hdr_i, colmap, store_cols = None, {}, []
        for i, row in enumerate(rows[:20]):
            vals = [norm(c) for c in row]
            if "produsent" in vals and ("produktnavn" in vals or "artikkelnr." in vals or "artikkelnr" in vals):
                for j, v in enumerate(vals):
                    if v in HEAD and HEAD[v] not in colmap.values():
                        colmap[j] = HEAD[v]
                    elif v.startswith("fordeling "):
                        store_cols.append(j)
                hdr_i = i
                break
        if hdr_i is None:
            continue
        for row in rows[hdr_i + 1:]:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            rec = {v: (row[j] if j < len(row) else None) for j, v in colmap.items()}
            producer = str(rec.get("producer") or "").strip()
            product = str(rec.get("product") or "").strip()
            if not producer and not product:
                continue                                        # separator / note row
            stores = sum(1 for j in store_cols if j < len(row) and to_num(row[j]))
            typ = rec.get("type")
            out.append({
                "art": str(rec.get("art") or "").strip() or None,
                "producer": producer or None,
                "product": product or None,
                "vintage": int(to_num(rec.get("vintage"))) if to_num(rec.get("vintage")) and 1900 < to_num(rec.get("vintage")) < 2100 else None,
                "country": str(rec.get("country") or "").strip() or None,
                "district": str(rec.get("district") or "").strip() or None,
                "municipality": str(rec.get("municipality") or "").strip() or None,
                "vineyard": str(rec.get("vineyard") or "").strip() or None,
                "classification": str(rec.get("classification") or "").strip() or None,
                "type_raw": str(typ or "").strip() or None,
                "type": canon_type(typ),
                "volume_l": to_num(rec.get("volume")),
                "abv": to_num(rec.get("abv")),
                "price": to_num(rec.get("price")),
                "importer": str(rec.get("importer") or "").strip() or None,
                "distributor": str(rec.get("distributor") or "").strip() or None,
                "qty": int(to_num(rec.get("qty"))) if to_num(rec.get("qty")) is not None else None,
                "value": to_num(rec.get("value")),
                "stores": stores or None,
                "period": period,
                "category_file": category,
            })
    wb.close()
    return out


def main():
    out = sys.argv[sys.argv.index("--out") + 1] if "--out" in sys.argv else "launch_history.json"
    files = sorted(glob.glob(os.path.join(DATA_DIR, "**", "*.xlsx"), recursive=True))
    all_recs, per_file = [], []
    for f in files:
        try:
            recs = parse_file(f)
        except Exception as e:
            per_file.append((os.path.basename(f), f"ERROR {e}"))
            continue
        all_recs.extend(recs)
        per_file.append((os.path.basename(f), f"{len(recs)} rows"))
    json.dump(all_recs, open(out, "w", encoding="utf-8"), ensure_ascii=False)
    print(f"{len(files)} files -> {len(all_recs)} launched products -> {out} ({os.path.getsize(out)//1024} KB)")
    # coverage
    def cov(field):
        return sum(1 for r in all_recs if r.get(field))
    for fld in ["producer", "product", "country", "district", "classification", "type",
                "price", "abv", "importer", "qty", "vintage"]:
        print(f"  {cov(fld):>6} / {len(all_recs)}  have {fld}")


if __name__ == "__main__":
    main()
