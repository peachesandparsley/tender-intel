"""
Bulk importer: validates a filled producer_upload_template.xlsx and merges
clean rows into wines.json as wine records.

Every row gets a verdict: OK, or a list of errors (rejected) / warnings
(imported, flagged). Nothing invalid ever silently enters the database.

Usage: python3 import_wines.py <filled_template.xlsx> [--dry-run]
"""
import json, re, sys
from openpyxl import load_workbook

COLOR_MAP = {"red": "red", "white": "white", "rosé": "rose", "rose": "rose",
             "sparkling white": "sparkling white", "sparkling rosé": "sparkling rose",
             "fortified": "fortified", "orange/skin-contact": "white"}
WOOD_MAP = {"none": "none", "old/neutral oak (discreet)": "old oak (discreet)",
            "barrel fermented": "barrel fermented", "new oak (prominent)": "prominent (new oak)",
            "partly oaked": "partly oaked (discreet)"}

def parse_grapes(text, errors):
    grapes = {}
    for part in re.split(r"[;,]\s*", str(text or "").strip()):
        if not part:
            continue
        m = re.match(r"(.+?)\s*[:=]\s*(\d+(?:\.\d+)?)\s*%?$", part.strip())
        if not m:
            errors.append(f"grape entry '{part}' not in 'variety:percent' form")
            continue
        grapes[m.group(1).strip().lower()] = float(m.group(2))
    total = sum(grapes.values())
    if grapes and abs(total - 100) > 0.5:
        errors.append(f"grape percentages sum to {total:g}, must be 100")
    return grapes

def parse_vintages(text, errors):
    t = str(text or "").strip().lower()
    if not t:
        return []
    if "nv" in t or "non-vintage" in t:
        return [0]
    years = [int(y) for y in re.findall(r"(19|20)\d{2}", str(text))] or \
            [int(y) for y in re.findall(r"\b((?:19|20)\d{2})\b", str(text))]
    years = [int(y) for y in re.findall(r"\b(19\d{2}|20\d{2})\b", str(text))]
    if not years:
        errors.append(f"could not read vintages from '{text}'")
    return years

def num(v, name, errors, required=False, lo=None, hi=None):
    if v is None or str(v).strip() == "":
        if required:
            errors.append(f"{name} is required")
        return None
    try:
        x = float(str(v).replace(",", "."))
    except ValueError:
        errors.append(f"{name}: '{v}' is not a number")
        return None
    if lo is not None and x < lo or hi is not None and x > hi:
        errors.append(f"{name}: {x:g} outside plausible range [{lo}–{hi}]")
        return None
    return x

def rep_status(flag, importer):
    f = str(flag or "").strip().lower()
    if f.startswith("yes"):
        return f"importer: {str(importer).strip()}" if str(importer or "").strip() else "represented (importer not stated)"
    if "open" in f:
        return "unrepresented"
    if f.startswith("no"):
        return "not seeking"
    return "unknown"

def import_file(path, dry=False):
    wb = load_workbook(path, data_only=True)
    if "Wines" not in wb.sheetnames:
        sys.exit("No 'Wines' sheet found — is this the right template?")
    ws = wb["Wines"]
    db = json.load(open("wines.json", encoding="utf-8"))
    existing = {(w["producer"].lower(), w["name"].lower()) for w in db["wines"]}
    next_id = max((int(w["id"][1:]) for w in db["wines"]), default=0) + 1

    report, added = [], 0
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        (producer, wine, country, region, appellation, colour, method, grapes_t,
         vintages_t, abv, sugar, wood, certs_t, cert_label, vine_age, macer,
         bottles, fob, rep_no, imp_no, rep_se, imp_se) = (list(row) + [None]*22)[:22]

        errors, warnings = [], []
        for name, v in [("Producer name", producer), ("Wine name", wine),
                        ("Country", country), ("Colour", colour)]:
            if not str(v or "").strip():
                errors.append(f"{name} is required")

        color = COLOR_MAP.get(str(colour or "").strip().lower())
        if str(colour or "").strip() and color is None:
            errors.append(f"Colour '{colour}' not recognized")
        if color and "sparkling" in color and not str(method or "").strip():
            warnings.append("sparkling wine without method — assumed unknown")
        if str(method or "").strip() and color and "sparkling" not in color:
            warnings.append("method given for a still wine — ignored")
            method = None

        grapes = parse_grapes(grapes_t, errors)
        if not grapes:
            errors.append("Grapes & % is required")
        vintages = parse_vintages(vintages_t, errors)
        if not vintages:
            errors.append("Vintages available is required")
        abv_v = num(abv, "ABV", errors, required=True, lo=4, hi=25)
        sugar_v = num(sugar, "Residual sugar", warnings, lo=0, hi=400)
        bottles_v = num(bottles, "Bottles available", errors, required=True, lo=1, hi=10_000_000)
        fob_v = num(fob, "Ex-cellar price EUR", errors, required=True, lo=0.5, hi=500)
        wood_v = WOOD_MAP.get(str(wood or "").strip().lower())
        if str(wood or "").strip() and wood_v is None:
            warnings.append(f"Wood '{wood}' not recognized — stored as free text")
            wood_v = str(wood).strip()
        vine_v = num(vine_age, "Old-vine age", warnings, lo=3, hi=200)
        macer_v = num(macer, "Maceration days", warnings, lo=0, hi=365)

        key = (str(producer or "").strip().lower(), str(wine or "").strip().lower())
        if key in existing:
            errors.append("duplicate: this producer+wine already exists in the database")

        if errors:
            report.append({"row": i, "wine": f"{producer} — {wine}", "status": "REJECTED",
                           "errors": errors, "warnings": warnings})
            continue

        rec = {
            "id": f"w{next_id:03d}",
            "producer": str(producer).strip(), "name": str(wine).strip(),
            "country": str(country).strip(), "region": str(region or "").strip(),
            "appellation": str(appellation or "").strip(),
            "grapes": grapes,
            "method": str(method).strip() if str(method or "").strip() else None,
            "vintages_available": vintages, "abv": abv_v,
            "sugar_g_l": sugar_v, "wood": wood_v,
            "certs": [c.strip() for c in re.split(r"[;,]", str(certs_t or "")) if c.strip()],
            "cert_on_label": str(cert_label or "").strip().lower() == "yes",
            "vines_age": vine_v, "maceration_days": macer_v,
            "volume_bottles": int(bottles_v), "fob_eur": fob_v,
            "representation": {"NO": rep_status(rep_no, imp_no),
                               "SE": rep_status(rep_se, imp_se)},
            "color": color,
            "source": f"bulk upload: {path.split('/')[-1]} row {i}",
        }
        db["wines"].append(rec)
        existing.add(key)
        next_id += 1
        added += 1
        report.append({"row": i, "wine": f"{producer} — {wine}", "status": "OK",
                       "errors": [], "warnings": warnings})

    if not dry and added:
        json.dump(db, open("wines.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    print(f"{'DRY RUN — ' if dry else ''}{added} wines imported, "
          f"{sum(1 for r in report if r['status']=='REJECTED')} rejected, "
          f"{len(db['wines'])} total in database\n")
    for r in report:
        print(f"  row {r['row']:>3} [{r['status']:^8}] {r['wine']}")
        for e in r["errors"]:
            print(f"           ERROR: {e}")
        for w in r["warnings"]:
            print(f"           warn:  {w}")
    return report

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    import_file(sys.argv[1], dry="--dry-run" in sys.argv)
