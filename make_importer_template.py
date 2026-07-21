"""Builds importer_portfolio_template.xlsx — identical to the producer template's
first 18 columns (no representation columns; it's the importer's own book).
Lighter bar: only producer, wine, country, colour required — the platform
enriches and the importer completes over time."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFFFF0")
EX_FILL = PatternFill("solid", fgColor="EDEDED")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Border(*[Side(style="thin", color="C0C0C0")] * 4)

COLS = [
    ("Producer name*", 28), ("Wine name*", 30), ("Country*", 16), ("Region", 18),
    ("Appellation / WO / AOC", 24), ("Colour*", 16), ("Sparkling method", 26),
    ("Grapes & % (var:pct; ...)", 30), ("Vintages available", 18), ("ABV %", 10),
    ("Residual sugar g/l", 16), ("Wood / oak", 22), ("Certifications", 30),
    ("Certification on label?", 20), ("Old-vine age (years)", 18), ("Maceration days", 16),
    ("Bottles available", 16), ("Ex-cellar price EUR", 18),
]
EXAMPLE = ["Kanonkop Estate", "Kadette Pinotage", "South Africa", "Stellenbosch",
           "WO Stellenbosch", "Red", "", "pinotage:100", "2023; 2024", 14.0, 2.5,
           "Old/neutral oak (discreet)", "WIETA", "no", "", "", 60000, 4.20]

wb = Workbook()
ins = wb.active
ins.title = "Instructions"
ins.column_dimensions["A"].width = 100
for i, (txt, size, bold) in enumerate([
    ("Importer portfolio upload — instructions", 14, True),
    ("", 10, False),
    ("One row per wine you represent. Only Producer, Wine, Country and Colour are required to", 10, False),
    ("start — but every extra field filled means more tenders matched automatically.", 10, False),
    ("A blank field shows as 'needs confirmation' instead of a clean match, so completing your", 10, False),
    ("rows (from the producer's tech sheet) is directly worth money at tender time.", 10, False),
    ("", 10, False),
    ("Grapes & %: 'variety:percent' with semicolons, summing to 100 (e.g. syrah:60; grenache:40).", 10, False),
    ("Vintages: 2023; 2024 — or NV. Prices: producer ex-cellar per 75cl, EUR.", 10, False),
    ("The grey row is an example — start on the row below it.", 10, False),
], 1):
    c = ins.cell(row=i, column=1, value=txt)
    c.font = Font(name=ARIAL, size=size, bold=bold)

ws = wb.create_sheet("Wines")
ws.freeze_panes = "A2"
for j, (name, width) in enumerate(COLS, 1):
    c = ws.cell(row=1, column=j, value=name)
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = THIN
    ws.column_dimensions[get_column_letter(j)].width = width
ws.row_dimensions[1].height = 30
for j, v in enumerate(EXAMPLE, 1):
    c = ws.cell(row=2, column=j, value=v)
    c.font = Font(name=ARIAL, size=10, italic=True)
    c.fill = EX_FILL
    c.border = THIN
for r in range(3, 303):
    for j in range(1, len(COLS) + 1):
        c = ws.cell(row=r, column=j)
        c.font = Font(name=ARIAL, size=10)
        c.fill = INPUT_FILL
        c.border = THIN

def dv(formula, cols):
    d = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(d)
    for col in cols:
        d.add(f"{col}3:{col}302")

dv('"Red,White,Rosé,Sparkling white,Sparkling rosé,Fortified,Orange/skin-contact"', ["F"])
dv('"Traditional method / Cap Classique,Charmat / tank,Pét-nat / ancestral,Carbonated"', ["G"])
dv('"None,Old/neutral oak (discreet),Barrel fermented,New oak (prominent),Partly oaked"', ["L"])
dv('"yes,no"', ["N"])

lg = ws.cell(row=1, column=len(COLS) + 2, value="Fill the pale-yellow cells. Grey row = example. * = required.")
lg.font = Font(name=ARIAL, size=9, italic=True)
lg.fill = YELLOW
wb.save("importer_portfolio_template.xlsx")
print("importer_portfolio_template.xlsx written")
