"""Builds producer_upload_template.xlsx — the bulk-upload sheet producers fill in.
One row per wine. Data-validation dropdowns for enums, an Instructions sheet,
and one realistic example row."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFFFF0")
EX_FILL = PatternFill("solid", fgColor="EDEDED")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Border(*[Side(style="thin", color="C0C0C0")] * 4)

COLS = [
    ("Producer name*",            28, "Legal/trading name, e.g. Simonsig Estate"),
    ("Wine name*",                30, "Cuvée name as on label"),
    ("Country*",                  16, "Country of origin"),
    ("Region",                    18, "e.g. Stellenbosch, Wachau"),
    ("Appellation / WO / AOC",    24, "Formal origin designation on label"),
    ("Colour*",                   16, "Dropdown"),
    ("Sparkling method",          26, "Dropdown — leave blank for still wine"),
    ("Grapes & % (var:pct; ...)*",30, "e.g. chardonnay:60; pinot noir:40 — must sum to 100"),
    ("Vintages available*",       18, "e.g. 2022; 2023 — write NV for non-vintage"),
    ("ABV %*",                    10, "e.g. 13.5"),
    ("Residual sugar g/l",        16, "Total incl. dosage for sparkling"),
    ("Wood / oak*",               22, "Dropdown"),
    ("Certifications",            30, "e.g. Organic (EU); Fairtrade; WIETA"),
    ("Certification on label?",   20, "Dropdown yes/no"),
    ("Old-vine age (years)",      18, "Oldest-parcel age if relevant, else blank"),
    ("Maceration days",           16, "Skin contact, if relevant"),
    ("Bottles available*",        16, "Volume you can commit for one campaign"),
    ("Ex-cellar price EUR*",      18, "FOB per 75cl bottle, EUR"),
    ("Represented in Norway?",    22, "Dropdown"),
    ("Norwegian importer (if any)",26, "Name, or blank"),
    ("Represented in Sweden?",    22, "Dropdown"),
    ("Swedish importer (if any)", 26, "Name, or blank"),
]

EXAMPLE = ["Simonsig Estate", "Kaapse Vonkel Brut Blanc de Blancs", "South Africa",
           "Stellenbosch", "WO Stellenbosch", "Sparkling white",
           "Traditional method / Cap Classique", "chardonnay:100", "2021; 2022",
           12.0, 7.0, "None", "WIETA", "no", "", "", 12000, 6.50,
           "No — open to offers", "", "Yes", "Vinunic AB"]

wb = Workbook()

# --- Instructions sheet ---
ins = wb.active
ins.title = "Instructions"
ins.column_dimensions["A"].width = 100
lines = [
    ("Producer wine upload — instructions", 14, True),
    ("", 10, False),
    ("Fill in one row per wine on the 'Wines' sheet. Columns marked * are required.", 10, False),
    ("The grey row is an example — do not delete it, do not edit it; start on the row below.", 10, False),
    ("", 10, False),
    ("Why every field matters: Nordic monopoly tenders specify requirements clause by clause", 10, False),
    ("(grape %, sugar limits, wood regime, certifications on the label, volume minimums).", 10, False),
    ("A blank field means 'unknown' — your wine will show as 'needs confirmation' instead of", 10, False),
    ("matching outright. Complete rows match more tenders.", 10, False),
    ("", 10, False),
    ("Grapes & %: use 'variety:percent' separated by semicolons, summing to 100.", 10, False),
    ("   Example: chenin blanc:60; grenache blanc:40", 10, False),
    ("Vintages: semicolons between years (2022; 2023). Write NV for non-vintage.", 10, False),
    ("Prices: ex-cellar (FOB) per 75cl bottle in EUR. This is used only to pre-screen", 10, False),
    ("   which tender price bands your wine can realistically compete in.", 10, False),
    ("Representation: 'No — open to offers' makes your wine visible to importers as an", 10, False),
    ("   introduction opportunity when it matches a live tender.", 10, False),
    ("", 10, False),
    ("Save the file and upload it — every row is validated and you get a per-row report.", 10, False),
]
for i, (txt, size, bold) in enumerate(lines, 1):
    c = ins.cell(row=i, column=1, value=txt)
    c.font = Font(name=ARIAL, size=size, bold=bold)

# --- Wines sheet ---
ws = wb.create_sheet("Wines")
ws.freeze_panes = "A2"
for j, (name, width, tip) in enumerate(COLS, 1):
    c = ws.cell(row=1, column=j, value=name)
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = THIN
    ws.column_dimensions[get_column_letter(j)].width = width
ws.row_dimensions[1].height = 30

# example row (grey)
for j, v in enumerate(EXAMPLE, 1):
    c = ws.cell(row=2, column=j, value=v)
    c.font = Font(name=ARIAL, size=10, italic=True)
    c.fill = EX_FILL
    c.border = THIN

# input rows styling
for r in range(3, 203):
    for j in range(1, len(COLS) + 1):
        c = ws.cell(row=r, column=j)
        c.font = Font(name=ARIAL, size=10)
        c.fill = INPUT_FILL
        c.border = THIN

def dv(formula, cols, allow_blank=True):
    d = DataValidation(type="list", formula1=formula, allow_blank=allow_blank,
                       showDropDown=False)
    ws.add_data_validation(d)
    for col in cols:
        d.add(f"{col}3:{col}202")

dv('"Red,White,Rosé,Sparkling white,Sparkling rosé,Fortified,Orange/skin-contact"', ["F"])
dv('"Traditional method / Cap Classique,Charmat / tank,Pét-nat / ancestral,Carbonated"', ["G"])
dv('"None,Old/neutral oak (discreet),Barrel fermented,New oak (prominent),Partly oaked"', ["L"])
dv('"yes,no"', ["N"])
dv('"Yes,No — open to offers,No — not seeking"', ["S", "U"])

# legend
lg = ws.cell(row=1, column=len(COLS) + 2,
             value="Fill the pale-yellow cells. Grey row = example. * = required.")
lg.font = Font(name=ARIAL, size=9, italic=True)
lg.fill = YELLOW

wb.save("producer_upload_template.xlsx")
print("producer_upload_template.xlsx written")
