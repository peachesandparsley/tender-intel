"""
Parser for Vinmonopolet lanseringsplan / tender-list Excel files (v2 —
hardened against the real 2020-1 English edition; keeps heuristics so
Norwegian editions and layout drift still parse).

Usage:
    python3 parse_lanseringsplan.py <file.xlsx> [-o specs.json]
"""
import json, re, sys, unicodedata
from openpyxl import load_workbook

# canonical field -> aliases (normalized). Order within list irrelevant;
# matching is two-pass: exact first, then longest-substring.
ALIASES = {
    "launch":     ["launch", "lansering", "lanseringsdato", "launch date", "slippdato", "dato"],
    "ref":        ["reference", "referanse", "ref", "anbudsnr", "tender", "varenummer"],
    "main_type":  ["main product type", "hovedvaregruppe", "hovedgruppe"],
    "group":      ["product type", "producy type", "varetype", "varegruppe", "product group", "produktgruppe", "category"],
    "country":    ["country", "land", "opprinnelse", "origin"],
    "region":     ["region", "omrade", "distrikt", "district"],
    "subregion":  ["sub region", "subregion", "underomrade"],
    "appellation":["quality/appellation", "appellation", "appelation", "kvalitet/klassifisering", "klassifisering"],
    "sensory":    ["sensory criteria", "sensoriske kriterier", "sensorisk"],
    "addspec":    ["additional specifications", "tilleggsspesifikasjoner", "andre krav"],
    "spec":       ["specifications", "specification", "spesifikasjon", "spesifikasjoner",
                   "beskrivelse", "description", "krav", "produktspesifikasjon"],
    "vintage":    ["vintage", "argang"],
    "packaging":  ["packaging", "emballasje"],
    "unit":       ["unit size", "volum", "storrelse", "flaskestorrelse"],
    "price_text": ["retail price", "prisomrade", "price range", "prisklasse", "veil. pris", "pris", "price"],
    "offer":      ["type of offer", "tilbudstype"],
    "deadline":   ["deadline", "frist", "tilbudsfrist", "offer deadline"],
    "quality":    ["quality criteria", "kvalitetskriterier"],
    "selection":  ["range", "utvalg", "assortment", "delutvalg", "basis/parti"],
}

def norm(s):
    if s is None:
        return ""
    s = str(s).translate(str.maketrans({"ø": "o", "Ø": "O", "æ": "ae", "Æ": "Ae"}))
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()

def map_headers(row_vals):
    cells = [(i, norm(c)) for i, c in enumerate(row_vals) if norm(c)]
    mapping = {}
    # pass 1: exact
    for i, n in cells:
        for field, names in ALIASES.items():
            if field not in mapping and n in names and i not in mapping.values():
                mapping[field] = i
                break
    # pass 2: substring, longest alias wins
    for i, n in cells:
        if i in mapping.values():
            continue
        best = None
        for field, names in ALIASES.items():
            if field in mapping:
                continue
            for a in names:
                if a in n and (best is None or len(a) > best[1]):
                    best = (field, len(a))
        if best:
            mapping[best[0]] = i
    return mapping

def parse_price_range(text):
    """'<300' / 'kr 130-179,90' / 'under 200' / '130 - 180' -> (lo, hi) NOK."""
    t = norm(text).replace(",", ".")
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", t)]
    if not nums:
        return None, None
    if t.startswith("<") or "under" in t or "inntil" in t or "max" in t or "opp til" in t:
        return None, nums[0]
    if t.startswith(">") or "over" in t or ("fra" in t and len(nums) == 1):
        return nums[0], None
    if len(nums) == 1:
        return nums[0] * 0.9, nums[0] * 1.1
    return min(nums), max(nums)

def parse_litres(text):
    """'75 cl' / '0,75 l' / '375 ml' -> litres (default 0.75)."""
    t = norm(text).replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)\s*(cl|ml|l|liter|litre)", t)
    if not m:
        return 0.75
    v, u = float(m.group(1)), m.group(2)
    return v / 100 if u == "cl" else v / 1000 if u == "ml" else v

def parse_workbook(path):
    wb = load_workbook(path, data_only=True)
    specs = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_i, mapping = None, {}
        for i, row in enumerate(rows[:40]):
            m = map_headers(row)
            if "group" in m and ("spec" in m or "price_text" in m) and len(m) >= 4:
                header_i, mapping = i, m
                break
        if header_i is None:
            continue
        for row in rows[header_i + 1:]:
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            def g(f):
                i = mapping.get(f)
                v = row[i] if i is not None and i < len(row) else None
                return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""
            if not g("group") and not g("spec"):
                continue
            # 2026-2+ format: requirement clauses live in "Quality criteria" +
            # "Additional specifications"; "Sensory criteria" is the tasting focus.
            spec_text = g("spec")
            quality = g("quality")
            if "sensory" in mapping:
                spec_text = " ".join(x for x in [spec_text, g("quality"), g("addspec"), g("packaging")] if x)
                quality = g("sensory")
            lo, hi = parse_price_range(g("price_text"))
            # old format: merged "Quality Criteria" often spans two columns
            if "sensory" not in mapping:
                qi = mapping.get("quality")
                if qi is not None and qi + 1 < len(row) and row[qi + 1] and (qi + 1) not in mapping.values():
                    quality = f"{quality}, {str(row[qi+1]).strip()}".strip(", ")
            deadline = g("deadline").split(" ")[0]  # drop 00:00:00
            specs.append({
                "sheet": ws.title,
                "ref": g("ref"),
                "launch": g("launch"),
                "selection": g("selection"),
                "main_type": g("main_type"),
                "group": g("group"),
                "country": g("country"),
                "region": " / ".join(x for x in [g("region"), g("subregion")] if x),
                "appellation": g("appellation"),
                "spec": spec_text,
                "vintage": g("vintage"),
                "packaging": g("packaging"),
                "litres": parse_litres(g("unit")),
                "price_text": g("price_text"),
                "price_lo": lo,
                "price_hi": hi,
                "offer": g("offer"),
                "deadline": deadline,
                "quality": quality,
            })
    return specs

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    out = sys.argv[sys.argv.index("-o") + 1] if "-o" in sys.argv else "specs.json"
    specs = parse_workbook(sys.argv[1])
    with open(out, "w", encoding="utf-8") as f:
        json.dump(specs, f, ensure_ascii=False, indent=1)
    from collections import Counter
    print(f"parsed {len(specs)} specs -> {out}")
    print("countries:", dict(Counter(s['country'] for s in specs).most_common(12)))
